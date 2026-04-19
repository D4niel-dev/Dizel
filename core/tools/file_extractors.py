"""
file_extractors.py — Extract text from various file formats.

Each extractor returns raw text. Cleaning and truncation happen later.
Dependencies are imported lazily so the app doesn't crash if a library
is missing — it just returns a helpful error.
"""

import os
import csv
import json
import io


def extract_text(file_path: str) -> tuple:
    """
    Dispatch to the right extractor based on extension.

    Returns (text, file_type, metadata_dict).
    Raises ValueError for unsupported types.
    """
    ext = os.path.splitext(file_path)[1].lower()

    extractors = {
        ".pdf":   _extract_pdf,
        ".docx":  _extract_docx,
        ".doc":   _extract_docx,
        ".txt":   _extract_plaintext,
        ".md":    _extract_plaintext,
        ".json":  _extract_json,
        ".jsonl": _extract_jsonl,
        ".csv":   _extract_csv,
        ".tsv":   _extract_tsv,
        ".xlsx":  _extract_xlsx,
        ".xls":   _extract_xlsx,
    }

    # Code files — treat as plaintext with syntax label
    code_exts = {
        ".py", ".js", ".ts", ".jsx", ".tsx",
        ".html", ".css", ".scss",
        ".c", ".cpp", ".h", ".hpp",
        ".java", ".go", ".rs", ".rb",
        ".sh", ".bat", ".ps1",
        ".yaml", ".yml", ".toml", ".ini", ".cfg",
        ".xml", ".sql",
    }

    if ext in extractors:
        return extractors[ext](file_path)

    if ext in code_exts:
        return _extract_code(file_path, ext)

    raise ValueError(f"No extractor for: {ext}")


def _extract_pdf(path: str) -> tuple:
    """Extract text from PDF using PyMuPDF."""
    try:
        import fitz
    except ImportError:
        raise ImportError(
            "PyMuPDF is required for PDF extraction. "
            "Install it with: pip install PyMuPDF"
        )

    doc = fitz.open(path)
    pages = []
    for i, page in enumerate(doc):
        text = page.get_text("text")
        if text.strip():
            pages.append(text)
    doc.close()

    meta = {"pages": len(pages)}
    return "\n\n".join(pages), "pdf", meta


def _extract_docx(path: str) -> tuple:
    """Extract text from DOCX using python-docx."""
    try:
        import docx
    except ImportError:
        raise ImportError(
            "python-docx is required for DOCX extraction. "
            "Install it with: pip install python-docx"
        )

    doc = docx.Document(path)
    paragraphs = [p.text for p in doc.paragraphs if p.text.strip()]
    meta = {"paragraphs": len(paragraphs)}
    return "\n\n".join(paragraphs), "docx", meta


def _extract_plaintext(path: str) -> tuple:
    """Read plain text or markdown files."""
    encodings = ["utf-8", "utf-8-sig", "latin-1", "cp1252"]
    for enc in encodings:
        try:
            with open(path, "r", encoding=enc) as f:
                text = f.read()
            ext = os.path.splitext(path)[1].lstrip(".")
            meta = {"lines": text.count("\n") + 1}
            return text, ext or "txt", meta
        except (UnicodeDecodeError, UnicodeError):
            continue
    raise ValueError(f"Cannot decode file: {path}")


def _extract_json(path: str) -> tuple:
    """Pretty-print a JSON file."""
    with open(path, "r", encoding="utf-8") as f:
        data = json.load(f)
    text = json.dumps(data, indent=2, ensure_ascii=False)
    meta = {"type": type(data).__name__}
    return text, "json", meta


def _extract_jsonl(path: str) -> tuple:
    """Read JSONL, show first N entries."""
    max_entries = 20
    entries = []
    total = 0
    with open(path, "r", encoding="utf-8") as f:
        for line in f:
            total += 1
            if len(entries) < max_entries:
                try:
                    obj = json.loads(line.strip())
                    entries.append(json.dumps(obj, ensure_ascii=False))
                except json.JSONDecodeError:
                    entries.append(line.strip())

    text = "\n".join(entries)
    if total > max_entries:
        text += f"\n\n[... {total - max_entries} more entries not shown ...]"
    meta = {"total_entries": total, "shown": min(total, max_entries)}
    return text, "jsonl", meta


def _extract_csv(path: str) -> tuple:
    """Read CSV and format as a readable table."""
    return _extract_delimited(path, ",", "csv")


def _extract_tsv(path: str) -> tuple:
    """Read TSV and format as a readable table."""
    return _extract_delimited(path, "\t", "tsv")


def _extract_delimited(path: str, delimiter: str, ftype: str) -> tuple:
    """Generic delimited file reader."""
    max_rows = 50
    rows = []
    with open(path, "r", encoding="utf-8", newline="") as f:
        reader = csv.reader(f, delimiter=delimiter)
        for i, row in enumerate(reader):
            if i >= max_rows + 1:  # +1 for header
                break
            rows.append(row)

    if not rows:
        return "(empty file)", ftype, {"rows": 0}

    # Count total rows
    with open(path, "r", encoding="utf-8") as f:
        total_rows = sum(1 for _ in f) - 1  # minus header

    # Format as table
    header = rows[0]
    lines = [" | ".join(header)]
    lines.append(" | ".join("---" for _ in header))
    for row in rows[1:]:
        lines.append(" | ".join(row))

    text = "\n".join(lines)
    if total_rows > max_rows:
        text += f"\n\n[... {total_rows - max_rows} more rows not shown ...]"

    meta = {"columns": len(header), "total_rows": total_rows, "shown": min(total_rows, max_rows)}
    return text, ftype, meta


def _extract_xlsx(path: str) -> tuple:
    """Extract text from Excel files using openpyxl."""
    try:
        import openpyxl
    except ImportError:
        raise ImportError(
            "openpyxl is required for XLSX extraction. "
            "Install it with: pip install openpyxl"
        )

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sheets = []
    max_rows_per_sheet = 30

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i >= max_rows_per_sheet + 1:
                break
            cells = [str(c) if c is not None else "" for c in row]
            rows.append(" | ".join(cells))

        if rows:
            header = f"[Sheet: {sheet_name}]"
            sheets.append(f"{header}\n" + "\n".join(rows))

    wb.close()
    text = "\n\n".join(sheets)
    meta = {"sheets": len(sheets)}
    return text, "xlsx", meta


def _extract_code(path: str, ext: str) -> tuple:
    """Read source code files with language annotation."""
    lang = ext.lstrip(".")
    with open(path, "r", encoding="utf-8") as f:
        code = f.read()

    lines = code.count("\n") + 1
    text = f"```{lang}\n{code}\n```"
    meta = {"language": lang, "lines": lines}
    return text, f"code ({lang})", meta
